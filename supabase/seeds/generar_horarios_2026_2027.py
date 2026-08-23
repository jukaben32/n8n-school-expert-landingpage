# -*- coding: utf-8 -*-
"""Genera el SQL para cargar los horarios 2026-2027 en produccion.

Lee el Excel CORREGIDO directamente (no el volcado de texto: hay celdas con
saltos de linea dentro, como "Lengua Espanola\\nCaligrafia", que al volcarlas
a texto plano se parten en dos filas).
"""
import openpyxl

SCHOOL = '0001da6e-2fe8-4dc9-97bf-8eadb7ee944e'
XLSX = r"C:\Users\IA Power Engine\Downloads\Horarios_20262027_MentorIApp_corregido.xlsx"

# ---------------------------------------------------------------- docentes --
# Nombre en el horario -> id en `staff`. Los nombres difieren del registro
# oficial (tildes, apellidos de mas): se mapean a mano, uno por uno.
STAFF = {
    # Secundaria
    'Yucleidy Pozo Pio': '5a0a9bbc-5ac5-4e0c-a7e9-05f82f835266',
    'Jennifer Sánchez': 'a58fe2dc-5829-482c-9334-8376e5b45bfe',
    'Alejandrina Lake': 'ad45e58b-9a0f-4a01-b17a-224ac6402ab2',
    'Marcelis Santos': '7e93f0b7-8340-44b9-b20e-2275318b477b',
    'Yendry Paulino': '532753e7-ecb4-4168-bea8-e8dcfc4834b0',
    'Ruth Mingo': '15124231-2a22-4e79-933b-ce3671dd5f1c',
    'Diarkis León': '263f34b8-673d-40ad-b827-203ee2a6a7da',
    'Juana Alcántara': 'cc833d1d-cbce-4182-b6be-ce2a9714bd52',
    'Jennifer Liliana Soriano': '062d636e-cf17-44fc-b09d-1bb5d6d12564',
    'Génesis Rodríguez': '0d5ed330-5733-49d2-a20c-2a60189decdd',
    # Orlando no existe todavia en `staff`; se crea en este mismo script.
    'Orlando Natera': 'b7c1e5a2-9d34-4f68-8a10-3c5e7f2b6d91',
    # Primaria (titulares)
    'Vianela Santana': '6aeaf7f5-b9f3-46c5-b2dc-cee91da758f5',
    'Indira Peguero': '9aec518a-4c6f-4fd7-bf3b-f09574be7333',
    'Manuela Vicente': 'c46118c5-e38d-43f4-b907-1c69aed4c9e7',
    'Teresa Santana': '21f58700-5951-49f6-90b7-18f3de6b6759',
    'Juana Iris Severino': '2a74c703-2d99-48d5-979f-dffbd3e07631',
    # Primaria (area de Ingles / Amco)
    'Maríanelis Calderón': '47e2602b-0544-4d73-85c9-9482b8db5c18',
    'Yuleimi Lugo Ochoa': '8943000e-41d5-447a-9480-427cb5e23c94',
}
ORLANDO_ID = STAFF['Orlando Natera']
SIN_ASIGNAR = '(sin asignar en el documento)'

# ---------------------------------------------------------------- materias --
# Variantes del documento -> nombre canonico. Sin esto quedarian materias
# duplicadas ("Ingles"/"Inglés", "Naturales"/"Ciencias Naturales", ...).
SUBJECTS = {
    'Lengua Española': 'Lengua Española',
    'Lengua Española / Caligrafía': 'Lengua Española / Caligrafía',
    'Leng. Española / Caligrafía': 'Lengua Española / Caligrafía',
    'Lengua Española\nCaligrafía': 'Lengua Española / Caligrafía',
    'Caligrafía': 'Caligrafía',
    'Matemática': 'Matemática',
    'Ciencias Naturales': 'Ciencias Naturales',
    'Naturales': 'Ciencias Naturales',
    'Biología': 'Biología',
    'Física': 'Física',
    'Ciencias Sociales': 'Ciencias Sociales',
    'Inglés': 'Inglés',
    'Ingles': 'Inglés',
    'Formación Humana': 'Formación Humana',
    'Educación Artística': 'Educación Artística',
    'Artística': 'Educación Artística',
    'Educación Física': 'Educación Física',
    'Educación Física (Práctica)': 'Educación Física',
    'Orientación Educativa': 'Orientación Educativa',
    'Orientación': 'Orientación Educativa',
    'Salida Optativa': 'Salida Optativa',
}

DIAS = {'Lunes': 1, 'Martes': 2, 'Miércoles': 3, 'Jueves': 4, 'Viernes': 5}
ORDINALES = {'1ro': '1ro.', '2do': '2do.', '3ro': '3ro.', '4to': '4to.',
             '5to': '5to.', '6to': '6to.'}

# Franjas de secundaria: no existen todavia, se crean en este script.
PERIODOS_SEC = [
    ('Sec. Bloque 1', '07:30', '08:20', 10, 'P1'),
    ('Sec. Bloque 2', '08:20', '09:10', 11, 'P2'),
    ('Sec. Bloque 3', '09:10', '10:00', 12, 'P3'),
    ('Sec. Bloque 4', '10:00', '10:50', 13, 'P4'),
    ('Sec. Recreo',   '10:50', '11:10', 14, None),
    ('Sec. Bloque 5', '11:10', '12:10', 15, 'P5'),
    ('Sec. Bloque 6', '12:10', '13:00', 16, 'P6'),
]
SEC_POR_PID = {pid: nombre for nombre, _, _, _, pid in PERIODOS_SEC if pid}

# Franjas de primaria: ya existen en la base, se referencian por nombre.
PRI_POR_HORA = {
    '7:40 – 8:30': 'Bloque 1',
    '8:30 – 9:20': 'Bloque 2',
    '10:20 – 11:10': 'Bloque 3',
    '11:10 – 11:50': 'Bloque 4',
    '11:50 – 12:30': 'Bloque 5',
}


def limpiar_docente(valor):
    """Quita las anotaciones del Excel: '(supuesto: docente titular)' etc."""
    if not valor:
        return None
    nombre = valor.split('(supuesto')[0].split('(asignado')[0].strip()
    if not nombre or nombre.startswith(SIN_ASIGNAR):
        return None
    return nombre


def sq(texto):
    """Escapa comillas simples para SQL."""
    return texto.replace("'", "''")


wb = openpyxl.load_workbook(XLSX)
filas, avisos = [], []

for hoja, nivel in (('Secundaria', 'Secundaria'), ('Primaria', 'Primaria')):
    ws = wb[hoja]
    for r in range(2, ws.max_row + 1):
        grado = ws.cell(r, 1).value
        dia = ws.cell(r, 2).value
        periodo = ws.cell(r, 3).value
        materia = ws.cell(r, 4).value
        docentes = ws.cell(r, 5).value
        if not grado or not dia or not periodo or not materia:
            continue

        grado, dia = str(grado).strip(), str(dia).strip()
        periodo, materia = str(periodo).strip(), str(materia).strip()

        if grado not in ORDINALES or dia not in DIAS:
            avisos.append(f'fila {hoja}!{r}: grado/dia no reconocido ({grado}, {dia})')
            continue

        canonica = SUBJECTS.get(materia)
        if not canonica:
            avisos.append(f'fila {hoja}!{r}: materia sin mapear -> {materia!r}')
            continue

        if nivel == 'Secundaria':
            periodo_nombre = SEC_POR_PID.get(periodo.split()[0])
        else:
            periodo_nombre = PRI_POR_HORA.get(periodo)
        if not periodo_nombre:
            avisos.append(f'fila {hoja}!{r}: franja sin mapear -> {periodo!r}')
            continue

        # Franja compartida de 1ro: los dos profesores de Ingles la reclaman.
        # Por la estructura del area de Ingles (Amco), 1er ciclo de secundaria
        # (1ro, 2do, 3ro) es de Orlando Natera; Yendry cubre 2do ciclo (4to-6to).
        nombres = [limpiar_docente(x) for x in str(docentes or '').split(';')]
        nombres = [n for n in nombres if n]
        if len(nombres) > 1:
            nombres = ['Orlando Natera'] if 'Orlando Natera' in nombres else nombres[:1]
            avisos.append(f'fila {hoja}!{r}: franja compartida resuelta a {nombres[0]}')

        staff_id = None
        if nombres:
            staff_id = STAFF.get(nombres[0])
            if not staff_id:
                avisos.append(f'fila {hoja}!{r}: docente sin mapear -> {nombres[0]!r}')

        filas.append({
            'grade': f'{ORDINALES[grado]} {nivel}',
            'day': DIAS[dia],
            'periodo': periodo_nombre,
            'materia': canonica,
            'staff': staff_id,
        })

# ----------------------------------------------------------------- salida ---
usadas = sorted({f['materia'] for f in filas})
out = []
w = out.append

w('-- Carga de horarios 2026-2027 -- generado automaticamente.')
w('-- Fuente: Horarios_20262027_MentorIApp_corregido.xlsx')
w('begin;')
w('')
w('-- 1. Orlando Antoine Natera: docente de Ingles de 1er ciclo de secundaria')
w('--    (area de Ingles/Amco). No existia en `staff`.')
w(f"""insert into staff (id, school_id, first_name, last_name, role, specialty)
values ('{ORLANDO_ID}', '{SCHOOL}', 'Orlando Antoine', 'Natera', 'teacher',
        'Inglés (Nivel secundario) primer ciclo')
on conflict (id) do nothing;""")
w('')
w('-- 2. Catalogo de materias (la tabla estaba vacia).')
for m in usadas:
    w(f"""insert into subjects (school_id, name) select '{SCHOOL}', '{sq(m)}'
where not exists (select 1 from subjects where school_id='{SCHOOL}' and name='{sq(m)}');""")
w('')
w('-- 3. Las 7 franjas que ya existian son las de primaria.')
w(f"update class_periods set level='primaria' where school_id='{SCHOOL}' and level is null;")
w('')
w('-- 4. Franjas de secundaria (no existian).')
for nombre, ini, fin, orden, _ in PERIODOS_SEC:
    w(f"""insert into class_periods (school_id, name, start_time, end_time, sort_order, level)
select '{SCHOOL}', '{nombre}', '{ini}', '{fin}', {orden}, 'secundaria'
where not exists (select 1 from class_periods where school_id='{SCHOOL}' and name='{nombre}');""")
w('')
w('-- 5. Normalizar grade_level: es el campo que une el horario con cada')
w('--    familia por RLS, asi que una variante de texto deja a esa familia')
w('--    sin ver su horario.')
for viejo, nuevo in [('6to Secundaria', '6to. Secundaria'),
                     ('1ro de Secundaria', '1ro. Secundaria'),
                     ('3r0. Primaria', '3ro. Primaria'),
                     ('4to. de Primaria', '4to. Primaria')]:
    w(f"update students set grade_level='{nuevo}' where school_id='{SCHOOL}' and grade_level='{viejo}';")
    w(f"update teacher_assignments set grade_level='{nuevo}' where school_id='{SCHOOL}' and grade_level='{viejo}';")
w('')
w(f'-- 6. El horario: {len(filas)} clases.')
for f in filas:
    staff = f"(select id from staff where id='{f['staff']}')" if f['staff'] else 'null'
    w(f"""insert into class_schedules (school_id, grade_level, day_of_week, period_id, subject_id, staff_id)
select '{SCHOOL}', '{f['grade']}', {f['day']},
       (select id from class_periods where school_id='{SCHOOL}' and name='{f['periodo']}'),
       (select id from subjects where school_id='{SCHOOL}' and name='{sq(f['materia'])}'),
       {staff}
on conflict (school_id, grade_level, day_of_week, period_id) do update
   set subject_id=excluded.subject_id, staff_id=excluded.staff_id, updated_at=now();""")
w('')
w('commit;')

with open('carga_horarios.sql', 'w', encoding='utf-8') as fh:
    fh.write('\n'.join(out))

print(f'Clases a cargar : {len(filas)}')
print(f'Materias        : {len(usadas)} -> {", ".join(usadas)}')
print(f'Grados          : {len(sorted({f["grade"] for f in filas}))}')
print(f'Sin docente     : {sum(1 for f in filas if not f["staff"])}')
print(f'\nAvisos ({len(avisos)}):')
for a in avisos[:20]:
    print('  -', a)
